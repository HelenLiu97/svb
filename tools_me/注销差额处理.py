"""
coding:utf-8
@software:PyCharm
@Time:2023/9/4 14:48
@Author:Helen
"""
import datetime
from openpyxl.reader.excel import load_workbook
from mysql_tools import SqlData


def main(user_id, n_time, cardnumber, refund_limit):

    before_balance = SqlData.search_user_field('balance', user_id)
    SqlData.update_balance(round(refund_limit, 2), user_id)
    balance = SqlData.search_user_field("balance", user_id)
    if refund_limit > 0:
        trans_type = '收入'
    else:
        trans_type = '支出'
    SqlData.insert_account_trans(n_time, trans_type, "注销差额", cardnumber,
                                 abs(refund_limit), before_balance, balance, user_id)
    print(cardnumber, refund_limit)


wb = load_workbook(r'C:\Users\Administrator\Documents\WeChat Files\wxid_gr0jmsweeht722\FileStorage\File\2023-11\注销差额恢复13.xlsx')
sheets = wb.worksheets  # 获取当前所有的sheet
# 获取第一张sheet
sheet1 = sheets[0]
rows = sheet1.max_row
# 获取工作表总列数
cols = sheet1.max_column
# 总行，总列
print(rows, cols)
# 读取
all_list = []
for r in range(1, rows + 1):
    row_list = []
    for i in range(1, cols + 1):
        cell_value = sheet1.cell(row=r, column=i).value
        row_list.append(cell_value)
    card_no, user_name, amount = row_list
    user_id = SqlData.search_user_field_name('id', user_name)
    n_time = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    print(user_id, n_time, card_no, amount)
    if not card_no or not user_id:
        print('数据异常')
        break
    main(user_id, n_time, card_no, amount)


if __name__ == "__main__":
    pass